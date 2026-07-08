"""Per-step evaluation against the PK gold set (docs/PPPK.xlsx, PK_Query sheet).

Where the count-based harness (harness.py) scores only the final result count, this
module scores each pipeline stage against its own column in the gold set, so
a regression can be traced to the stage that caused it (CONST-9). Each stage gets the
comparator that fits the shape of its output:

  * termite (Stage 0)   — set match over recognized entity labels
  * decompose (Stage 1) — routing-type-boolean: set match over (field, type) pairs
  * translate (Stage 2) — set F1 over the emitted field names
  * aggregate / machine query (Stage 3) — structural: top operator + matched fields

All comparators run offline against the gold cells.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field

from oppp.config import REPO_ROOT
from oppp.models import PipelineResult

PPPK_PATH = REPO_ROOT / "docs" / "PPPK.xlsx"
PERSTEP_SHEET = "PK_Query"

_TYPE_RE = re.compile(r"(\w+)\[(filter|question)\]")
_FIELD_EQ_RE = re.compile(r"\b([A-Za-z][A-Za-z0-9]+)\s*=")
_ENTITY_RE = re.compile(r"([A-Z_]+)\s*:\s*([^;()]+)")


@dataclass
class StepScore:
    step: str
    score: float
    expected: list[str] = field(default_factory=list)
    got: list[str] = field(default_factory=list)
    detail: str = ""


def load_perstep_cases() -> list[dict]:
    """Load PK gold cases from PPPK.xlsx PK_Query sheet.

    Returns rows as dicts with at least 'Query', 'Expected Count', and 'Quety number'.
    Uses openpyxl (optional); falls back to an empty list if not installed.
    """
    if not PPPK_PATH.exists():
        return []
    try:
        import openpyxl
    except ImportError:
        return []
    wb = openpyxl.load_workbook(PPPK_PATH, read_only=True, data_only=True)
    if PERSTEP_SHEET not in wb.sheetnames:
        return []
    ws = wb[PERSTEP_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        d = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        result.append(d)
    return result


def find_perstep_case(query_number: int) -> dict | None:
    for row in load_perstep_cases():
        num = row.get("Quety number") or row.get("Query number")
        try:
            if int(num) == query_number:
                return row
        except (TypeError, ValueError):
            pass
    return None


def _f1(expected: set[str], got: set[str]) -> float:
    if not expected and not got:
        return 1.0
    if not expected or not got:
        return 0.0
    tp = len(expected & got)
    if tp == 0:
        return 0.0
    precision = tp / len(got)
    recall = tp / len(expected)
    return 2 * precision * recall / (precision + recall)


def _entity_labels(cell: str) -> set[str]:
    return {
        m.group(2).strip().lower() for m in _ENTITY_RE.finditer(cell or "") if m.group(2).strip()
    }


def _routing_pairs(cell: str) -> set[tuple[str, str]]:
    return {(m.group(1), m.group(2)) for m in _TYPE_RE.finditer(cell or "")}


def _translate_fields(cell: str) -> set[str]:
    return {m.group(1) for m in _FIELD_EQ_RE.finditer(cell or "")}


def _match_fields(node) -> set[str]:
    """All constraint field names anywhere in a machine-query boolean tree."""
    out: set[str] = set()
    if isinstance(node, dict):
        for op, body in node.items():
            if op in ("MATCH", "REGEX", "RANGE", "DATE_RANGE", "EMPTY") and isinstance(body, dict):
                if "field" in body:
                    out.add(body["field"])
            else:
                out |= _match_fields(body)
    elif isinstance(node, list):
        for child in node:
            out |= _match_fields(child)
    return out


def _top_op(query: dict) -> str | None:
    return next(iter(query), None) if isinstance(query, dict) and query else None


def score_termite(result: PipelineResult, gold_cell: str) -> StepScore:
    expected = _entity_labels(gold_cell)
    got = {
        a.label.strip().lower() for a in (result.enhanced.annotations if result.enhanced else [])
    }
    return StepScore("termite", _f1(expected, got), sorted(expected), sorted(got))


def score_decompose(result: PipelineResult, gold_cell: str) -> StepScore:
    expected = _routing_pairs(gold_cell)
    got = {(c.field, c.type.value) for c in result.decomposition.components}
    exp_s = sorted(f"{a}[{b}]" for a, b in expected)
    got_s = sorted(f"{a}[{b}]" for a, b in got)
    return StepScore("decompose", _f1(expected, got), exp_s, got_s)


def score_translate(result: PipelineResult, gold_cell: str) -> StepScore:
    expected = _translate_fields(gold_cell)
    got = {sq.field for sq in result.subqueries}
    return StepScore("translate", _f1(expected, got), sorted(expected), sorted(got))


def score_machine_query(result: PipelineResult, gold_cell: str) -> StepScore:
    try:
        gold = json.loads(gold_cell)
    except (json.JSONDecodeError, TypeError):
        return StepScore("machine_query", 0.0, detail="unparseable gold machine-query cell")
    if result.machine_query is None:
        return StepScore("machine_query", 0.0, detail="no machine query produced")
    expected = _match_fields(gold.get("query", {}))
    got = _match_fields(result.machine_query.query)
    field_f1 = _f1(expected, got)
    top_match = _top_op(gold.get("query", {})) == _top_op(result.machine_query.query)
    score = field_f1 if top_match else field_f1 * 0.5
    return StepScore(
        "machine_query",
        score,
        sorted(expected),
        sorted(got),
        detail=f"top_op {'match' if top_match else 'differ'}; field_f1={field_f1:.2f}",
    )


_COMPARATORS = {
    "termite": score_termite,
    "decompose": score_decompose,
    "translate": score_translate,
    "machine query": score_machine_query,
}


def compare_steps(result: PipelineResult, gold_row: dict) -> dict[str, StepScore]:
    """Score every stage of one run against the gold row's per-stage columns."""
    scores: dict[str, StepScore] = {}
    for column, comparator in _COMPARATORS.items():
        scores[column] = comparator(result, str(gold_row.get(column, "") or ""))
    return scores
