"""Evaluation over the PK gold set (docs/PPPK.xlsx, PK_Query sheet).

The metric is **result-count accuracy**: each gold row carries the question
(``Query``) and an expected number of matching records (``Expected Count``). We
translate the question, execute the machine query against the PharmaPendium API,
read ``countTotal``, and compare it to the expected count.

Reported: how often the query is valid, how often it executes, exact-count match
rate, and within-tolerance rate (counts drift as the DB updates, so a tolerance
band is the realistic signal).
"""

from __future__ import annotations

import csv
import re
import time
from dataclasses import dataclass, field
from pathlib import Path

from oppp.config import REPO_ROOT, get_settings
from oppp.execute import execute_count
from oppp.pipeline import run_pipeline
from oppp.services.base import get_service

PPPK_PATH = REPO_ROOT / "docs" / "PPPK.xlsx"
PPPK_SHEET = "PK_Query"


def _parse_expected(cell) -> int | None:
    if cell is None:
        return None
    s = str(cell).strip()
    if not s:
        return None
    m = re.search(r"\d[\d,]*", s)
    return int(m.group(0).replace(",", "")) if m else None


@dataclass
class CaseResult:
    query_number: str
    question: str
    ok: bool  # produced a valid machine query
    expected: int | None = None
    actual: int | None = None
    executed: bool = False
    issues: list[str] = field(default_factory=list)
    exec_error: str | None = None

    @property
    def exact(self) -> bool:
        return (
            self.expected is not None and self.actual is not None and self.expected == self.actual
        )

    def within(self, tol: float) -> bool:
        if self.expected is None or self.actual is None:
            return False
        if self.expected == 0:
            return self.actual == 0
        return abs(self.actual - self.expected) / self.expected <= tol

    @property
    def ratio(self) -> float | None:
        if self.expected is None or self.actual is None or self.expected == 0:
            return None
        return self.actual / self.expected


@dataclass
class EvalReport:
    tolerance: float = 0.10
    cases: list[CaseResult] = field(default_factory=list)

    @property
    def valid_rate(self) -> float:
        return _mean([1.0 if c.ok else 0.0 for c in self.cases])

    @property
    def executed_rate(self) -> float:
        return _mean([1.0 if c.executed else 0.0 for c in self.cases])

    @property
    def scored(self) -> list[CaseResult]:
        return [c for c in self.cases if c.executed and c.expected is not None]

    @property
    def exact_match_rate(self) -> float:
        s = self.scored
        return _mean([1.0 if c.exact else 0.0 for c in s]) if s else 0.0

    @property
    def within_tol_rate(self) -> float:
        s = self.scored
        return _mean([1.0 if c.within(self.tolerance) else 0.0 for c in s]) if s else 0.0


def _mean(xs: list[float]) -> float:
    return sum(xs) / len(xs) if xs else 0.0


def load_pk_cases() -> list[dict]:
    """Load PK gold cases from PPPK.xlsx PK_Query sheet.

    Returns rows as dicts with keys matching the spreadsheet headers, including
    'Query', 'Expected Count', and 'Quety number' (sic — typo in the source file).
    Uses openpyxl; raises FileNotFoundError if the workbook is missing.
    """
    if not PPPK_PATH.exists():
        raise FileNotFoundError(
            f"PK gold set not found: {PPPK_PATH}. Place PPPK.xlsx in the docs/ directory."
        )
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError(
            "Reading PPPK.xlsx requires openpyxl: pip install openpyxl"
        ) from e
    wb = openpyxl.load_workbook(PPPK_PATH, read_only=True, data_only=True)
    if PPPK_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{PPPK_SHEET}' not found in PPPK.xlsx. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb[PPPK_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        d = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        result.append(d)
    return result


def load_gold_cases() -> list[dict]:
    """Alias for load_pk_cases() — returns the PK gold set rows as dicts.

    Kept for backward compatibility with existing callers that expect a list of
    dicts with a 'question' key. Maps 'Query' -> 'question' and
    'Expected Count' -> 's' for the per-field comparison view.
    """
    rows = load_pk_cases()
    out = []
    for r in rows:
        d = dict(r)
        if "Query" in d and "question" not in d:
            d["question"] = d["Query"]
        if "Expected Count" in d and "s" not in d:
            d["s"] = d["Expected Count"]
        out.append(d)
    return out


def evaluate(
    *,
    service: str = "pk",
    expander: str = "noop",
    enhancer: str = "noop",
    decomposer: str = "gazetteer",
    translator: str = "deterministic",
    aggregator: str = "deterministic",
    normalizer: str = "fuzzy",
    tolerance: float = 0.10,
    execute: bool = True,
    limit: int | None = None,
) -> EvalReport:
    """Run each PK gold question and compare executed ``countTotal`` to the expected count.

    Reads the question (``Query``) and expected count (``Expected Count``) from
    the PPPK.xlsx PK_Query sheet. Defaults to the offline doubles (gazetteer +
    deterministic) so the harness is cheap and hermetic; pass decomposer='llm'/
    aggregator='llm' to evaluate the production pipeline.
    ``execute=False`` skips the API call: only validity is measured.
    """
    svc = get_service(service)
    report = EvalReport(tolerance=tolerance)
    rows = load_pk_cases()
    if limit:
        rows = rows[:limit]

    for row in rows:
        q = str(row.get("Query") or "").strip()
        if not q:
            continue
        qnum = str(row.get("Quety number") or row.get("Query number") or "")
        result = run_pipeline(
            q,
            service,
            expander=expander,
            enhancer=enhancer,
            decomposer=decomposer,
            translator=translator,
            aggregator=aggregator,
            normalizer=normalizer,
            probe_open_filters=execute,
        )
        cr = CaseResult(
            query_number=qnum,
            question=q,
            ok=result.ok,
            expected=_parse_expected(row.get("Expected Count")),
            issues=[i.message for i in result.issues],
        )
        if execute and result.ok and result.machine_query is not None:
            ex = execute_count(result.machine_query, svc)
            cr.executed = ex.ok
            cr.actual = ex.count_total
            cr.exec_error = ex.error
        report.cases.append(cr)
    return report


# --- report export ----------------------------------------------------------
CASE_COLUMNS = [
    "query_number",
    "question",
    "expected",
    "actual",
    "ok",
    "executed",
    "exact",
    "within_tol",
    "ratio",
    "issues",
    "exec_error",
]


def _case_row(c: CaseResult, tol: float) -> list:
    return [
        c.query_number,
        c.question,
        c.expected,
        c.actual,
        c.ok,
        c.executed,
        c.exact,
        c.within(tol),
        round(c.ratio, 4) if c.ratio is not None else None,
        "; ".join(c.issues),
        c.exec_error or "",
    ]


def _metric_rows(report: EvalReport) -> list[tuple[str, object]]:
    return [
        ("generated", time.strftime("%Y-%m-%d %H:%M:%S")),
        ("cases", len(report.cases)),
        ("tolerance", report.tolerance),
        ("valid_rate", round(report.valid_rate, 4)),
        ("executed_rate", round(report.executed_rate, 4)),
        ("exact_match_rate", round(report.exact_match_rate, 4)),
        ("within_tol_rate", round(report.within_tol_rate, 4)),
    ]


_PARAMS_TITLE = "Parameters"
_SUMMARY_TITLE = "Summary"
_BOLD_FIRST_CELLS = {_PARAMS_TITLE, _SUMMARY_TITLE, CASE_COLUMNS[0]}


def _report_blocks(report: EvalReport, run_config: dict | None) -> list[list]:
    """The top-of-report rows: the run parameters, then the summary metrics."""
    blocks: list[list] = [[_PARAMS_TITLE]]
    for k, v in (run_config or {}).items():
        blocks.append([k, v])
    blocks.append([])
    blocks.append([_SUMMARY_TITLE])
    for k, v in _metric_rows(report):
        blocks.append([k, v])
    blocks.append([])
    return blocks


def write_report(report: EvalReport, path: str | Path, run_config: dict | None = None) -> Path:
    """Write the eval report to ``path``. Format from the extension (default ``.xlsx``).

    The report is a single sheet: the **parameters used** and the **run summary** on
    top, then the per-case table. ``.xlsx`` (the default — a bare path with no suffix
    becomes ``.xlsx``) needs the optional ``openpyxl`` dependency; ``.csv`` uses
    the stdlib. Raises ``ValueError`` on an unsupported extension.
    """
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == "":
        path = path.with_suffix(".xlsx")
        suffix = ".xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)

    header_blocks = _report_blocks(report, run_config)
    case_rows = [_case_row(c, report.tolerance) for c in report.cases]

    if suffix == ".csv":
        with open(path, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            for row in header_blocks:
                w.writerow(row)
            w.writerow(CASE_COLUMNS)
            for row in case_rows:
                w.writerow(row)
    elif suffix == ".xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as e:  # pragma: no cover - optional extra
            raise RuntimeError(
                "Excel output needs openpyxl: pip install openpyxl"
            ) from e
        wb = Workbook()
        ws = wb.active
        ws.title = "report"
        for row in header_blocks:
            ws.append(row)
        ws.append(CASE_COLUMNS)
        for row in case_rows:
            ws.append(row)
        for r in ws.iter_rows():
            if r and r[0].value in _BOLD_FIRST_CELLS:
                for cell in r:
                    if cell.value is not None:
                        cell.font = Font(bold=True)
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 48
        wb.save(path)
    else:
        raise ValueError(f"unsupported report extension {suffix!r}; use .xlsx or .csv")
    return path
